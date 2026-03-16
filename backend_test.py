#!/usr/bin/env python3
"""
Backend API testing for Excel Export functionality
Tests all 4 export endpoints: students, groups, schedules, graduates
"""

import os
import requests
import sys
from datetime import datetime
from io import BytesIO
import openpyxl

class ExcelExportTester:
    def __init__(self, base_url="https://ux-clarity-update.preview.emergentagent.com"):
        self.base_url = base_url
        self.token = None
        self.branch_id = "edfdd0c3-e0cf-4f7a-b1fc-8e2c925f0ae1"
        self.tests_run = 0
        self.tests_passed = 0
        self.headers = {'Content-Type': 'application/json'}

    def log(self, message):
        """Log test results"""
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")

    def run_test(self, name, method, endpoint, expected_status=200, data=None, check_excel=False):
        """Run a single API test"""
        url = f"{self.base_url}/{endpoint}"
        headers = self.headers.copy()
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'

        self.tests_run += 1
        self.log(f"🔍 Testing {name}...")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers, timeout=30)
            elif method == 'POST':
                response = requests.post(url, json=data, headers=headers, timeout=30)
            else:
                raise Exception(f"Unsupported method: {method}")

            success = response.status_code == expected_status
            
            if success and check_excel:
                # Validate Excel file
                content_type = response.headers.get('content-type', '')
                if 'spreadsheet' not in content_type and 'excel' not in content_type:
                    self.log(f"❌ Wrong content type: {content_type}")
                    return False, {}
                
                # Try to parse Excel
                try:
                    wb = openpyxl.load_workbook(BytesIO(response.content))
                    ws = wb.active
                    row_count = ws.max_row
                    col_count = ws.max_column
                    self.log(f"📊 Excel file: {row_count} rows, {col_count} columns")
                    
                    # Check first row contains headers
                    if row_count > 0:
                        headers_row = [cell.value for cell in ws[1]]
                        self.log(f"📋 Headers: {headers_row}")
                        if all(h for h in headers_row):
                            self.log("✅ Excel validation successful")
                        else:
                            self.log("⚠️  Some headers are empty")
                    
                except Exception as e:
                    self.log(f"❌ Excel parsing failed: {str(e)}")
                    return False, {}

            if success:
                self.tests_passed += 1
                self.log(f"✅ Passed - Status: {response.status_code}")
                if check_excel:
                    return True, {"content_length": len(response.content)}
            else:
                self.log(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                if response.text:
                    self.log(f"Response: {response.text[:200]}")

            return success, response.json() if not check_excel and response.content else {}

        except Exception as e:
            self.log(f"❌ Failed - Error: {str(e)}")
            return False, {}

    def login(self, username="admin", password="admin123"):
        """Login and get token"""
        success, response = self.run_test(
            "Admin Login",
            "POST",
            "api/auth/login",
            200,
            data={"username": username, "password": password}
        )
        if success and 'token' in response:
            self.token = response['token']
            self.log(f"✅ Login successful, role: {response.get('user', {}).get('role', 'unknown')}")
            return True
        self.log("❌ Login failed - no token received")
        return False

    def test_export_endpoint(self, export_type):
        """Test a specific export endpoint"""
        endpoint = f"api/admin/export/{export_type}?branch_id={self.branch_id}"
        success, response_data = self.run_test(
            f"Export {export_type.title()}",
            "GET",
            endpoint,
            200,
            check_excel=True
        )
        return success

    def test_health_check(self):
        """Test basic health endpoint"""
        return self.run_test("Health Check", "GET", "api/health", 200)[0]

    def run_all_tests(self):
        """Run all export tests"""
        self.log("🚀 Starting Excel Export API Tests")
        self.log(f"Backend URL: {self.base_url}")
        self.log(f"Branch ID: {self.branch_id}")
        
        # Test health
        if not self.test_health_check():
            self.log("❌ Health check failed - stopping tests")
            return False
        
        # Login
        if not self.login():
            self.log("❌ Login failed - stopping tests")
            return False
        
        # Test all export endpoints
        export_types = ["students", "groups", "schedules", "graduates"]
        for export_type in export_types:
            success = self.test_export_endpoint(export_type)
            if not success:
                self.log(f"❌ {export_type} export failed")

        # Summary
        self.log("\n" + "="*50)
        self.log(f"📊 TEST SUMMARY")
        self.log(f"Tests run: {self.tests_run}")
        self.log(f"Tests passed: {self.tests_passed}")
        self.log(f"Success rate: {(self.tests_passed/self.tests_run*100):.1f}%")
        
        if self.tests_passed == self.tests_run:
            self.log("🎉 All tests PASSED!")
            return True
        else:
            failed = self.tests_run - self.tests_passed
            self.log(f"❌ {failed} test(s) FAILED")
            return False

def main():
    """Main test runner"""
    tester = ExcelExportTester()
    success = tester.run_all_tests()
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())